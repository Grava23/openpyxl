#импортируем нужные нам фреймфорки
import pandas
from collections import defaultdict, Counter
import openpyxl

#открываем файл exel (лист1) и всю информацию добовляем в переменную
excel_data = pandas.read_excel('logs.xlsx', sheet_name='log') 
#преобразум его в список из словарей
excel_data_log = excel_data.to_dict(orient='records')

#список самых популярных браузеров, товаров, в том числе для мужчин, и для женщин
barausers_dict = defaultdict(int)
goods_dict = defaultdict(int)
goods_dict_for_men = defaultdict(int)
goods_dict_for_women = defaultdict(int)

#Словари с помощью которых мы будем записывать в exel файл популярные товары и браузеры
popular_browser = {}
popular_goods = {}

#Словари посещений и покупак в месяц
list_of_visits = {}
list_of_sales = {}

#количество посещений браузера, и покупок товаров в месяц
quantity = 1


#получить браузер
current_browser = ''
#получить товар
current_item = ''

for element in excel_data_log:
#Получаем браузеры, для вычесления самых популярных
    barausers_dict[element['Браузер']] +=1
#Товары   
    goods = element['Купленные товары']
    goods = goods.split(',')
     
#получаем номер месяца
    date = element['Дата посещения']
    date_str = str(date)
    date_notime = date_str.split()[0]
#Номер месяца
    number_of_month = date_notime.split('-')[1]

#Выяснение самых предпочтительных товаров длдя мужчин и женщин
    for current_item in goods:
        if not ('Ещё 2 варианта' in current_item) and not ('Ещё 3 варианта' in current_item):
            goods_dict[current_item] += 1
            if element['Пол'] == 'м':
                goods_dict_for_men[current_item] += 1
            else:
                goods_dict_for_women[current_item] += 1
#здесь мы создаём "двумерный словарь" с помощью которого мы будем заполнять покупки товаров в каждом месяце
        if current_item in popular_goods.keys():
            list_of_sales = popular_goods[current_item]
        else:
            list_of_sales = defaultdict(int)
        list_of_sales[number_of_month] += quantity
        popular_goods[current_item] = list_of_sales
#А здесь тоже самое для посещение браузеров
#Получаем браузер
    current_browser = element['Браузер']
#проверить есть ли текущий браузер в нашем словаре:
    if str(current_browser) in popular_browser.keys():
#если есть, то получить список продаж по месяцам для текущего браузера
        list_of_visits = popular_browser[str(current_browser)]   
#если нет, то создать
    else:
        list_of_visits = defaultdict(int)
#обновить значение в list_of_sales   
    list_of_visits[number_of_month] += quantity
#засунуть обновленный list_of_sales в итоговый словарь
    
    popular_browser[str(current_browser)] = list_of_visits
           
#Счётчик для браузеров
barausers_counter = Counter(barausers_dict)
#Счётчик для самых популярных товаров
goods_counter = Counter(goods_dict)

#Счётчик самых предпочтительных товаров, для мужчин
goods_counter_for_men = Counter(list(goods_dict_for_men))
#вычесляем длинну счётчика для мужчин
len_counter_for_men = len(goods_counter_for_men)

#Счётчик самых предпочтительных товаров, для женщин
goods_counter_for_women = Counter(list(goods_dict_for_women))
#вычесляем длинну счётчика для женщин
len_counter_for_women = len(goods_counter_for_women)

#Находим самые популярные браузеры
most_popular_browsers = barausers_counter.most_common()
#Находим самые популярные товары и удаляем лишнее
most_popular_goods = goods_counter.most_common(8)
for i in most_popular_goods:
    if i[0] == 'Ещё 2 варианта' or i[0] == 'Ещё 3 варианта':
        del most_popular_goods[most_popular_goods.index(i)]

#список 7 самых популярных браузеров
seven_popular_browsers = []
for part in barausers_counter.most_common(7):
    str(part).split(',')
    seven_popular_browsers.append(part[0])

#список 7 самых популярных товаров
seven_popular_items = []
for item in goods_counter.most_common(7):
    str(item).split(',')
    seven_popular_items.append(item[0])
    
#открываем лист наешего файла с помощью openpyexel и делаем его активным
logs = openpyxl.load_workbook(filename='report.xlsx')
logs.active = 0
sheet = logs.active

# заполняем покупки самых поулярных товаров по месяцам
quantity_items = 0
for row in range(19, 26):
#здесь мы получаем ключ ввиде какого-то браузера
    key = seven_popular_items[quantity_items]
    sheet.cell(row=row, column=1).value = key
# А здесь его посещение по месяцам
    key = popular_goods[key]
    number = 0
    for key, value in key.items():
        if  int(key) < 10:
            number = int(key[1])
            number = number + 2
        else:
            number = int(key) + 2
        sheet.cell(row = row, column = number ).value = value
    quantity_items += 1
quantity_barowsers = 0

#также делаем и для браузеров 
for row in range(5,12):
    key = seven_popular_browsers[quantity_barowsers]
    sheet.cell(row=row, column=1).value = key
    key = popular_browser[key]
    number = 0
    for key,value in key.items():
        if  int(key) < 10:
            number = int(key[1])
            number = number + 2
        else:
            number = int(key) + 2
        sheet.cell(row = row, column = number ).value = value
    quantity_barowsers += 1



#самый популярный товар для мужчин
sheet['B31'] = goods_counter_for_men.most_common(1)[0][0]
#Самый популярный товар для для женщин
sheet['B32'] = goods_counter_for_women.most_common(1)[0][0]
#самый Непопулярный товар для мужчин
sheet['B33'] = goods_counter_for_men.most_common()[:-(len_counter_for_men + 1):-1][0][0]
#самый Непопулярный товар для женщин
sheet['B34'] = goods_counter_for_women.most_common()[:-(len_counter_for_women + 1):-1][0][0]

logs.save(filename = 'report.xlsx')

